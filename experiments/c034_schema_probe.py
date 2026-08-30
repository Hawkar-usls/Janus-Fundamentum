#!/usr/bin/env python3
import hashlib, json, urllib.request
from pathlib import Path
import openpyxl

URL='https://data.caltech.edu/records/0yw13-j0441/files/SARS_CoV_2_extreme_differences_in_viral_loads.xlsx?download=1'
EXPECTED_MD5='6e6216d751c95b6afb6a7d0d96da6f1a'

with urllib.request.urlopen(URL, timeout=60) as r:
    data=r.read()
md5=hashlib.md5(data).hexdigest()
if md5!=EXPECTED_MD5:
    raise RuntimeError(f'md5 mismatch {md5}')
Path('tmp').mkdir(exist_ok=True)
path=Path('tmp/c034_source.xlsx'); path.write_bytes(data)
wb=openpyxl.load_workbook(path, read_only=True, data_only=False)
out={'md5':md5,'bytes':len(data),'sheets':[]}
for ws in wb.worksheets:
    # Headers only: first non-empty row; no data values emitted.
    first_nonempty=None
    max_scan=min(ws.max_row,25)
    for ridx,row in enumerate(ws.iter_rows(min_row=1,max_row=max_scan,values_only=True), start=1):
        vals=[None if v is None else str(v) for v in row]
        if any(v not in (None,'') for v in vals):
            first_nonempty={'row':ridx,'cells':vals}
            break
    out['sheets'].append({'title':ws.title,'max_row':ws.max_row,'max_column':ws.max_column,'first_nonempty_row':first_nonempty})
Path('out').mkdir(exist_ok=True)
Path('out/C034_SCHEMA.json').write_text(json.dumps(out,indent=2),encoding='utf-8')
print('C034_SCHEMA='+json.dumps(out,sort_keys=True))
