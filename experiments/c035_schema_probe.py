#!/usr/bin/env python3
import hashlib, json, urllib.request
from pathlib import Path
import openpyxl

URL='https://data.caltech.edu/records/krbp7-h5z17/files/20260529_EarlyTypeIIIFNLimitsViralReplication_UnderlyingData_Fig1-5_Deidentfied.xlsx?download=1'
EXPECTED_MD5='0aa6236431794c44815c9bfb3d6301d5'
req=urllib.request.Request(URL,headers={'User-Agent':'Mozilla/5.0 JANUS-C035-public-data-replay/1.0'})
with urllib.request.urlopen(req,timeout=60) as r:data=r.read()
md5=hashlib.md5(data).hexdigest()
if md5!=EXPECTED_MD5: raise RuntimeError(f'md5 mismatch {md5}')
Path('tmp').mkdir(exist_ok=True); p=Path('tmp/c035.xlsx'); p.write_bytes(data)
wb=openpyxl.load_workbook(p,read_only=True,data_only=False)
out={'md5':md5,'bytes':len(data),'sheets':[]}
for ws in wb.worksheets:
    first=None
    for ridx,row in enumerate(ws.iter_rows(min_row=1,max_row=min(ws.max_row,30),values_only=True),start=1):
        vals=[None if v is None else str(v) for v in row]
        if any(v not in (None,'') for v in vals):
            first={'row':ridx,'cells':vals}; break
    out['sheets'].append({'title':ws.title,'max_row':ws.max_row,'max_column':ws.max_column,'first_nonempty_row':first})
Path('out').mkdir(exist_ok=True); Path('out/C035_SCHEMA.json').write_text(json.dumps(out,indent=2),encoding='utf-8')
print('C035_SCHEMA='+json.dumps(out,sort_keys=True))
